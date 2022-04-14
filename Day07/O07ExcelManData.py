
from openpyxl import load_workbook

wb = load_workbook("FirstExcel.xlsx")

wb.active = wb['Players']

ws = wb.active

# A6 : E11      => col - A - 1, E = 5; row - 6 to 11

# for row in ws.iter_rows(min_row=6, min_col=1, max_row=11, max_col=5):
for row in ws.iter_rows(min_row=6, min_col=1, max_row=11):
    for cell in row:
        if cell.value == "Lara":
            cell.value = "brain lara".upper()

wb.save("FirstExcel.xlsx")
