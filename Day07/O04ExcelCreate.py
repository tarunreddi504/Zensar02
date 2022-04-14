
from openpyxl import Workbook
from datetime import datetime

wb= Workbook()

ws = wb.active
ws.title = "MySheet"

ws["B5"] = "Hello World"
ws["B6"] = 85000
ws["B7"] = datetime.now()

wb.save("FirstExcel.xlsx")